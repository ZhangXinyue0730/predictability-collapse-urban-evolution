#!/usr/bin/env python3
"""Build a compact Excel workbook from municipal-land audit CSV outputs."""
from __future__ import annotations

import os
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUT_DIR = Path(os.environ.get("MUNICIPAL_AUDIT_DIR", "outputs/municipal_audit")).resolve()
XLSX = OUT_DIR / "道路市政转入核验汇总.xlsx"

SHEETS = [
    ("各时期市政转入总量", "municipal_period_summary.csv"),
    ("近年城市时期排名_按数量", "municipal_recent_rank_by_count.csv"),
    ("近年城市时期排名_按占比", "municipal_recent_rank_by_share.csv"),
    ("近年城市合计排名", "municipal_recent_city_total.csv"),
    ("近年来源拆解_全量", "municipal_source_breakdown_all.csv"),
    ("建议GIS核验案例", "municipal_recommended_cases.csv"),
]


def main() -> None:
    with pd.ExcelWriter(XLSX, engine="openpyxl") as writer:
        for sheet_name, file_name in SHEETS:
            path = OUT_DIR / file_name
            df = pd.read_csv(path)
            if sheet_name == "近年来源拆解_全量":
                df = df[df["period"].isin(["2015->2020", "2020->2024"])]
            if "排名" in sheet_name:
                df = df.head(120)
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(XLSX)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            width = min(max(max(len(str(c.value)) if c.value is not None else 0 for c in col) + 2, 10), 38)
            ws.column_dimensions[letter].width = width
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=False)
                if isinstance(cell.value, float):
                    cell.number_format = "0.0000"
    wb.save(XLSX)
    print(XLSX)


if __name__ == "__main__":
    main()
