from pathlib import Path
import os
from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(os.environ.get("PROJECT_ROOT", ".")).resolve()
SRC = ROOT / "city_extract_national/national_analysis_dynamic_t1/79城市新统计口径完整汇总大表_接入年鉴经济人口.xlsx"
OUTDIR = ROOT / "city_extract_national/national_analysis_dynamic_t1"
OUT_XLSX = OUTDIR / "第4章4.1_八时期城市更新总体统计表.xlsx"
OUT_CSV = OUTDIR / "第4章4.1_八时期城市更新总体统计表.csv"


def num(value, default=0):
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return value
    text = str(value).replace(",", "").replace("%", "").strip()
    if not text:
        return default
    return float(text)


def main():
    wb_src = load_workbook(SRC, read_only=True, data_only=True)
    ws_src = wb_src["master_with_yearbook"]
    headers = [c.value for c in next(ws_src.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    agg = {}
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        period = row[idx["period"]]
        if not period:
            continue
        if period not in agg:
            agg[period] = {
                "year_t1": int(num(row[idx["year_t1"]])),
                "year_t2": int(num(row[idx["year_t2"]])),
                "city_count": 0,
                "samples": 0,
                "updates": 0,
                "update_area_km2": 0.0,
            }
        agg[period]["city_count"] += 1
        agg[period]["samples"] += int(num(row[idx["samples"]]))
        agg[period]["updates"] += int(num(row[idx["updates"]]))
        agg[period]["update_area_km2"] += float(num(row[idx["update_area_km2"]]))

    period_order = sorted(agg, key=lambda p: agg[p]["year_t1"])
    rows = []
    for period in period_order:
        a = agg[period]
        rate = a["updates"] / a["samples"] if a["samples"] else 0
        rows.append(
            [
                period,
                a["year_t1"],
                a["year_t2"],
                a["city_count"],
                a["samples"],
                a["updates"],
                rate,
                rate * 100,
                a["update_area_km2"],
                a["updates"] / a["city_count"] if a["city_count"] else 0,
                a["update_area_km2"] / a["city_count"] if a["city_count"] else 0,
            ]
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "4.1总体统计表"
    data_ws = wb.create_sheet("图表数据")

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="5B9BD5")
    light_fill = PatternFill("solid", fgColor="F7FBFF")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers_cn = [
        "时期",
        "前一年",
        "后一年",
        "城市数",
        "有效样本数",
        "更新数",
        "更新率",
        "更新率%",
        "更新面积 km²",
        "平均每城更新数",
        "平均每城更新面积 km²",
    ]

    ws.merge_cells("A1:K1")
    ws["A1"] = "第4.1节 八时期城市更新总体统计表（79城市，新统计口径）"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:K2")
    ws["A2"] = (
        "口径：按 period 汇总 master_with_yearbook；更新率=Σ更新数/Σ有效样本数；"
        "更新面积来自 update_area_km2 汇总。更新仅统计前一期建成区范围内 1-6 类用地之间的变化，"
        "0 空地/未利用不计入更新。"
    )
    ws["A2"].font = Font(size=10, color="666666")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 36

    for col, h in enumerate(headers_cn, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

        data_ws.cell(row=1, column=col, value=h)
        data_ws.cell(row=1, column=col).font = Font(bold=True)

    for r, row in enumerate(rows, 5):
        for c, value in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if (r - 5) % 2 == 1:
                cell.fill = light_fill
            data_ws.cell(row=r - 3, column=c, value=value)

    for r in range(5, 5 + len(rows)):
        for col in (5, 6):
            ws.cell(r, col).number_format = "#,##0"
        ws.cell(r, 7).number_format = "0.00%"
        ws.cell(r, 8).number_format = "0.0"
        ws.cell(r, 9).number_format = "#,##0.0"
        ws.cell(r, 10).number_format = "#,##0.0"
        ws.cell(r, 11).number_format = "#,##0.0"

    widths = [16, 10, 10, 10, 16, 14, 12, 12, 16, 18, 22]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
        data_ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:K{4 + len(rows)}"

    cats = Reference(ws, min_col=1, min_row=5, max_row=4 + len(rows))
    bar = BarChart()
    bar.type = "col"
    bar.title = "八时期城市更新总体趋势"
    bar.y_axis.title = "更新数"
    bar.x_axis.title = "时期"
    bar.height = 8
    bar.width = 18
    bar.add_data(Reference(ws, min_col=6, max_col=6, min_row=4, max_row=4 + len(rows)), titles_from_data=True)
    bar.set_categories(cats)
    bar.legend.position = "b"

    line = LineChart()
    line.add_data(Reference(ws, min_col=8, max_col=8, min_row=4, max_row=4 + len(rows)), titles_from_data=True)
    line.set_categories(cats)
    line.y_axis.axId = 200
    line.y_axis.title = "更新率%"
    line.y_axis.crosses = "max"
    line.series[0].graphicalProperties.line.solidFill = "ED7D31"
    line.series[0].marker.symbol = "circle"
    line.series[0].marker.size = 6
    bar += line
    ws.add_chart(bar, "A15")

    ws.merge_cells("A31:K31")
    ws["A31"] = f"数据来源：{SRC.name} / master_with_yearbook；生成文件：{OUT_XLSX.name}"
    ws["A31"].font = Font(size=9, color="666666")

    import csv

    with open(OUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers_cn)
        writer.writerows(rows)

    wb.save(OUT_XLSX)

    print(OUT_XLSX)
    print(OUT_CSV)
    for row in rows:
        print(row[0], row[4], row[5], f"{row[7]:.2f}", f"{row[8]:.1f}")


if __name__ == "__main__":
    main()
