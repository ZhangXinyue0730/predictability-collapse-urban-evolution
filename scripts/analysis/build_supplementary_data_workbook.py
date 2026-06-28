from __future__ import annotations

import math
import os
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE = Path(os.environ.get("SUPPLEMENTARY_WORKBOOK_DIR", "outputs/supplementary_data_workbook")).resolve()
INTERMEDIATE = BASE / "intermediate"
OUT = BASE / "Supplementary_Data.xlsx"

SHEETS = [
    ("SD1_Master_79x8", "79-city by 8-period master table"),
    ("SD2_Multimodel_AUC_AP", "Three-model AUC/AP matrix"),
    ("SD3_Ablation_M0_M7", "M0-M7 ablation results"),
    ("SD4_Permutation_Importance", "Complete permutation importance results"),
    ("SD5_Transitions_All", "Complete land-use transition matrix"),
    ("SD6_Update_Expansion_Area", "Update and expansion area by city-period"),
    ("SD7_Regional_Tests", "Regional Kruskal-Wallis tests"),
    ("SD8_Syntax_Summary", "Street-network and syntax metrics"),
    ("SD9_Feature_Dictionary_167", "Dictionary of 167 model features"),
    ("SD10_Source_Index", "Source index for all sheets"),
]


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
TEXT_ALIGN = Alignment(vertical="top", wrap_text=False)


def clean_value(value):
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def styled_header(ws, value):
    cell = WriteOnlyCell(ws, value=value)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = HEADER_ALIGN
    return cell


def data_cell(ws, value):
    cell = WriteOnlyCell(ws, value=clean_value(value))
    cell.alignment = TEXT_ALIGN
    return cell


def set_column_widths(ws, columns: list[str]) -> None:
    for idx, name in enumerate(columns, start=1):
        width = min(max(len(str(name)) + 2, 12), 32)
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_sheet(wb: Workbook, sheet_name: str, description: str) -> tuple[str, int, int]:
    csv_path = INTERMEDIATE / f"{sheet_name}.csv"
    if not csv_path.exists():
        raise FileNotFoundError(csv_path)

    df = pd.read_csv(csv_path)
    ws = wb.create_sheet(title=sheet_name)
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    # A concise worksheet-level note survives round-trips and helps reviewers.
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.append([styled_header(ws, col) for col in df.columns])
    set_column_widths(ws, list(df.columns))

    for row in df.itertuples(index=False, name=None):
        ws.append([data_cell(ws, value) for value in row])

    return sheet_name, len(df), len(df.columns)


def main() -> None:
    wb = Workbook(write_only=True)
    summary_rows = []
    for sheet_name, description in SHEETS:
        summary_rows.append(write_sheet(wb, sheet_name, description))

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)

    print(f"wrote={OUT}")
    for sheet_name, rows, cols in summary_rows:
        print(f"{sheet_name}: rows={rows}, columns={cols}")


if __name__ == "__main__":
    main()
